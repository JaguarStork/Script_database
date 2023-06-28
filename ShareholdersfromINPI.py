import openpyxl
import pandas as pd
import requests

# Set the batch size
batch_size = 10000


#source excel   
# "C:/Users/kamal/Downloads/chiffres-cles-2021_filter_final.csv"
source_chiffreclesCSV =r"C:\Users\kamal\Downloads\Base_1.csv"
# result_chiffreclesXLSX="https://docs.google.com/spreadsheets/d/1QSzDIMU5M4iQbBhtmAp_0JGA--KygkqG/edit?usp=share_link&ouid=111156806361311190632&rtpof=true&sd=true"
result_chiffreclesXLSX =r"C:\Users\kamal\Downloads\Base_1.xlsx"
rolesfile=r"C:\Users\kamal\Downloads\INPI_RNE_2020_Dictionnaire_de_donnees.xlsx"

# URLs
urlLogin = "https://registre-national-entreprises.inpi.fr/api/sso/login"
inpi = "https://registre-national-entreprises.inpi.fr/api/companies/"

# #initiate result Xlsx 
# workbook = openpyxl.load_workbook(result_chiffreclesXLSX)
# worksheet = workbook['stakeholders']
# # Delete all columns
# worksheet.delete_cols(1, worksheet.max_column)
# worksheet.append("nom","prenoms","role")
# Login
data = {
    "username": "",
    "password": ""
}

response = requests.post(urlLogin, json=data)
access_token = response.json()['token']

headers = {
    "Authorization": "Bearer " + access_token
}

# Read the Excel file into a pandas dataframe in chunks
chunks = pd.read_csv(source_chiffreclesCSV,dtype=str,chunksize=batch_size)

# Create an empty list to store the stakeholders data
stakeholders_data = []

# # open Excel file and select the sheet with role codes
# wb = openpyxl.load_workbook('file.xlsx')
# sheet = wb['role codes']
# # extract role codes from the column 'codes'
# role_codes = [str(cell.value) for cell in sheet['codes']]

wb = openpyxl.load_workbook(rolesfile)
sheet = wb['role']
# extract role codes from the column 'codes' where stakeholder is set to yes
role_codes = []


for row in sheet.iter_rows(min_row=2, values_only=True):
    code, _, stakeholder = row[:3]
    if stakeholder == 'yes':
        role_codes.append(str(code))

# print ("role_codes ", role_codes)
# Loop through each chunk of the dataframe
nom_prenoms = []
for chunk in chunks:
    # Filter the chunk based on the value of the "filter" column
    filtered_chunk = chunk.loc[chunk['filter'] == 'ok']

    # Select the cells in the "siren" column
    siren_cells = filtered_chunk['Siren']

    # Loop through each value in the "siren" column
    count =1
    for siren in siren_cells:
        # Make a request with the value as a parameter in the URL of the REST webservice

        url = f"https://registre-national-entreprises.inpi.fr/api/companies/{siren}"
        response = requests.get(url,headers=headers)
        try:
            pouvoirs = response.json(
    )['formality']['content']['personneMorale']['composition']['pouvoirs']
        except KeyError:
    # handle the case where the key is missing
            pouvoirs = []
       
# loop through the `pouvoirs` list to find the object with `role = 71`
        for pouvoir in pouvoirs:
            # print(pouvoir)
            individu = pouvoir.get('individu')
            # if 'descriptionPersonne' in individu and 'role' in individu['descriptionPersonne']:
            # role = str(individu['descriptionPersonne']['role'])
            # if role in role_codes:
            if individu and individu.get('descriptionPersonne', {}).get('role') in role_codes:
            # extract the `nom` and `prenoms` from the object
                nom = individu['descriptionPersonne']['nom']
                prenoms = ",".join(individu['descriptionPersonne']['prenoms'])
                role = individu['descriptionPersonne']['role']
                # print(individu)
                # print("nom ",nom)
                # print("prenoms ",prenoms)
                # print("role ", role)
                nom_prenoms.append((siren,nom, prenoms,role))
        # print("noms prenoms roles")
        # print(nom_prenoms)
        print("Siren "+ str(siren) +" "+str(count))
        count =count +1

# workbook.save(result_chiffreclesXLSX)


# Create a dataframe from the stakeholders data
stakeholders_df = pd.DataFrame(nom_prenoms, columns=['Siren',
                               'Nom', 'Prenoms', 'Code'])

print(stakeholders_df)

# Write the stakeholders dataframe to a new tab called "stakeholders" in the same Excel file
with pd.ExcelWriter(result_chiffreclesXLSX,encoding='utf8',engine='openpyxl', mode='a',if_sheet_exists='replace' ) as writer:
    stakeholders_df.to_excel(writer, sheet_name='stakeholders', index=False)
